# modules/excel_writer.py

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from .utils import get_month_name, get_bl_akhir, get_bl_awal, get_rptag_addition


def create_excel(success_data, failed_data, periods, output_path, file_data_map):
    wb = Workbook()

    # Sheet 1: Sukses
    if success_data:
        ws_success = wb.active
        ws_success.title = "Sukses"
        month_headers = [get_month_name(period) for period in periods]
        headers = ["ID Pelanggan", "Nama Lengkap", "Tarif/Daya", "Jumlah Periode"] + month_headers + \
            ["Tagihan", "Denda", "Biaya Admin", "Total Tagihan", "Tambahan", "MarkUp", "Sumber File"]
        ws_success.append(headers)

        for record in success_data:
            idpel = record.get("customer_number", "")
            additional_data = file_data_map.get(idpel, {})

            # Menggabungkan 'segmentation' dan 'DAYA' untuk kolom "Tarif/Daya" tanpa .0
            segmentation = record.get('segmentation', '')
            daya = additional_data.get('DAYA', '')
            if isinstance(daya, float):
                daya = int(daya) if daya.is_integer() else daya  # Menghilangkan .0 jika integer
            tarif_daya = f"{segmentation} / {daya}"

            row = [
                record.get("customer_number", ""),
                record.get("customer_name", ""),
                tarif_daya,  # Tarif/Daya yang telah digabungkan tanpa .0
                len(record.get("bills", []))
            ]
            for period in periods:
                bill = next((bill for bill in record.get("bills", []) if bill.get("bill_period") == period), {})
                row.append(bill.get("amount", 0))
            tagihan = sum(bill.get("amount", 0) for bill in record.get("bills", []))
            denda = record.get("penalty_fee", 0)
            biaya_admin = record.get("admin_charge", 0)
            total_tagihan = tagihan + denda + biaya_admin
            tambahan = record.get("tambahan", 0)
            markup = tagihan + tambahan
            row.extend([tagihan, denda, biaya_admin, total_tagihan, tambahan, markup, record.get("source_file", "")])
            ws_success.append(row)

        # Format header
        for col_num, column_title in enumerate(headers, 1):
            cell = ws_success.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ws_success.column_dimensions[get_column_letter(col_num)].width = 20

        # Format isi data
        for row in ws_success.iter_rows(min_row=2, min_col=1, max_col=len(headers), max_row=ws_success.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                if isinstance(cell.value, int) or isinstance(cell.value, float):
                    cell.number_format = '#,##0'

    # Sheet 2: Gagal
    if failed_data:
        ws_failed = wb.create_sheet(title="Gagal")
        headers_failed = ["ID Pelanggan", "Error", "Sumber File"]
        ws_failed.append(headers_failed)
        for record in failed_data:
            ws_failed.append([record.get("customer_number", ""), record.get(
                "error", ""), record.get("source_file", "")])

        for col_num in range(1, len(headers_failed) + 1):
            cell = ws_failed.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ws_failed.column_dimensions[get_column_letter(col_num)].width = 30

        # Format isi data gagal
        for row in ws_failed.iter_rows(min_row=2, min_col=1, max_col=len(headers_failed), max_row=ws_failed.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # Sheet 3: TUL
    ws_tul = wb.create_sheet(title="TUL")
    tul_headers = ["NO", "IDPEL", "NO RBM", "NAMA GARDU", "NAMA PELANGGAN", "ALAMAT",
                   "GOL", "TRF", "DAYA", "BL Awal", "BL Akhir", "LBR", "RPTAG", "RPBK", "Sumber File"]
    ws_tul.append(tul_headers)

    for idx, record in enumerate(success_data, 1):
        idpel = record.get("customer_number", "")
        additional_data = file_data_map.get(idpel, {})
        lbr_value = len(record.get("bills", []))

        # Mendapatkan BL Awal berdasarkan kategori LBR
        bl_awal = get_bl_awal(lbr_value)

        # Mendapatkan BL Akhir (sama untuk semua baris)
        bl_akhir = get_bl_akhir()

        # Mendapatkan tambahan RPTAG berdasarkan source file
        source_file = record.get("source_file", "")
        rptag_addition = get_rptag_addition(source_file)

        # Mengedit nilai RPTAG dengan menambahkan nilai tambahan
        rptag_current = sum(bill.get("amount", 0) for bill in record.get("bills", []))
        rptag_new = rptag_current + rptag_addition

        row = [
            idx,  # NO
            idpel,  # IDPEL
            additional_data.get("NO RBM", ""),  # NO RBM
            additional_data.get("NAMA GARDU", ""),  # NAMA GARDU
            additional_data.get("NAMA PELANGGAN", ""),  # NAMA PELANGGAN
            additional_data.get("ALAMAT", ""),  # ALAMAT
            additional_data.get("GOL", ""),  # GOL
            additional_data.get("TRF", ""),  # TRF
            additional_data.get("DAYA", ""),  # DAYA
            bl_awal,  # BL Awal berdasarkan kategori LBR
            bl_akhir,  # BL Akhir
            f"({lbr_value}",  # LBR hanya tambahkan "(" di awal
            rptag_new,  # RPTAG setelah ditambah
            record.get("penalty_fee", 0),  # RPBK
            source_file  # Sumber File
        ]
        ws_tul.append(row)

    # Format header sheet TUL
    for col_num, column_title in enumerate(tul_headers, 1):
        cell = ws_tul.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws_tul.column_dimensions[get_column_letter(col_num)].width = 20

    # Format isi data sheet TUL
    for row in ws_tul.iter_rows(min_row=2, min_col=1, max_col=len(tul_headers), max_row=ws_tul.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                cell.number_format = '#,##0'

    wb.save(output_path)
    logging.info(f"\nHasil telah disimpan ke {output_path}")
