import logging
import os
from openpyxl import load_workbook


def load_customer_numbers_from_folder(folder_path):
    file_paths = []
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
        for file_name in os.listdir(folder_path):
            if file_name.endswith('.txt') or file_name.endswith('.xlsx'):
                file_paths.append(os.path.join(folder_path, file_name))
    else:
        logging.warning(f"Folder {folder_path} tidak ditemukan atau bukan folder.")

    return load_customer_numbers(file_paths)


def load_customer_numbers(file_paths):
    customer_numbers = []
    customer_sources = []
    for file_path in file_paths:
        if file_path.endswith('.txt'):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    numbers = [line.strip() for line in f if line.strip() and not line.startswith('#')]
                    customer_numbers.extend(numbers)
                    customer_sources.extend([file_path] * len(numbers))
                logging.info(f"Loaded {len(numbers)} customer numbers from {os.path.basename(file_path)}.")
            except Exception as e:
                logging.error(f"Error saat memuat customer numbers dari {os.path.basename(file_path)}: {e}")
        elif file_path.endswith('.xlsx'):
            customer_data = load_customer_numbers_xlsx(file_path)
            for idpel, data in customer_data.items():
                customer_numbers.append(idpel)
                customer_sources.append(file_path)
            logging.info(f"Loaded {len(customer_data)} customer numbers from {os.path.basename(file_path)}.")
    # Menghapus duplikasi
    seen = set()
    unique_customer_numbers = []
    unique_customer_sources = []
    for num, src in zip(customer_numbers, customer_sources):
        if num not in seen:
            seen.add(num)
            unique_customer_numbers.append(num)
            unique_customer_sources.append(src)
    logging.info(f"Total {len(unique_customer_numbers)} unique customer numbers loaded from {len(file_paths)} files.")
    return unique_customer_numbers, unique_customer_sources


def load_customer_numbers_xlsx(file_path):
    customer_data = {}  # Menggunakan dictionary dengan IDPEL sebagai key
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active  # Mengambil sheet aktif
        header = [cell.value for cell in ws[1]]  # Baris pertama dianggap sebagai header

        # Validasi kolom yang diperlukan
        required_columns = ["IDPEL", "NO RBM", "NAMA GARDU", "NAMA PELANGGAN", "ALAMAT", "GOL", "TRF", "DAYA"]
        if not all(col in header for col in required_columns):
            logging.warning(f"File {os.path.basename(file_path)} tidak memiliki semua kolom yang diperlukan.")
            return {}

        # Mendapatkan indeks kolom yang sesuai
        col_indices = {col: header.index(col) for col in required_columns}

        # Memuat data
        for row in ws.iter_rows(min_row=2, values_only=True):  # Mulai dari baris kedua
            idpel = str(row[col_indices["IDPEL"]]).strip() if row[col_indices["IDPEL"]] else None
            if idpel:
                customer_data[idpel] = {
                    "NO RBM": row[col_indices["NO RBM"]],
                    "NAMA GARDU": row[col_indices["NAMA GARDU"]],
                    "NAMA PELANGGAN": row[col_indices["NAMA PELANGGAN"]],
                    "ALAMAT": row[col_indices["ALAMAT"]],
                    "GOL": row[col_indices["GOL"]],
                    "TRF": row[col_indices["TRF"]],
                    "DAYA": row[col_indices["DAYA"]]
                }
        logging.info(f"Loaded data for {len(customer_data)} customers from {os.path.basename(file_path)}.")
        wb.close()
    except Exception as e:
        logging.error(f"Error saat memuat data dari {os.path.basename(file_path)}: {e}")

    return customer_data


def collect_tambahan_inputs(source_files):
    tambahan_dict = {}
    print("\n=== Input Nilai 'Tambahan' untuk Setiap Sumber File ===")
    unique_sources = set(source_files)
    for source in unique_sources:
        source_file_name = os.path.basename(source)
        while True:
            try:
                value = float(input(f"Masukkan nilai 'Tambahan' untuk sumber file '{source_file_name}': "))
                tambahan_dict[source_file_name] = value
                break
            except ValueError:
                print("Input tidak valid. Silakan masukkan angka.")
    print("========================================================\n")
    return tambahan_dict


def load_customer_numbers_from_files(file_path):
    """
    Fungsi ini memuat customer numbers dari satu file saja.

    Parameters:
        file_path (str): Path lengkap ke file (.txt atau .xlsx).

    Returns:
        tuple: (customer_numbers, customer_sources)
    """
    return load_customer_numbers([file_path])
